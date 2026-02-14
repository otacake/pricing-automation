from __future__ import annotations  # 型注釈の前方参照を許可し、循環参照を避けるため

"""
Output helpers for profit test results.
"""

from pathlib import Path  # パスの操作をOS非依存で行うため
import json  # JSON出力に使うため

from openpyxl import Workbook  # Excelファイル出力に使うため

from .config import load_optimization_settings, loading_surplus_threshold  # 最適化設定と閾値計算に使うため
from .diagnostics import build_run_summary  # 診断サマリ出力に使うため
from .optimize import OptimizationResult  # 最適化結果の型を参照するため
from .profit_test import ProfitTestBatchResult, ProfitTestResult, model_point_label  # 収益性検証結果の型と表示用に使うため


def _write_cashflow_sheet(ws, result: ProfitTestResult) -> None:  # キャッシュフローの詳細をExcelに書く
    headers = list(result.cashflow.columns)  # DataFrameの列名を取得する
    header_row = 4  # ヘッダーの開始行を固定する
    data_row_start = header_row + 1  # データの開始行を決める
    for col_idx, name in enumerate(headers, start=1):  # 列名を順番に書き込む
        ws.cell(row=header_row, column=col_idx, value=name)  # ヘッダー行に列名をセットする

    for row_offset, row in enumerate(result.cashflow.itertuples(index=False), start=0):  # 各行の値を書き込む
        for col_idx, value in enumerate(row, start=1):  # 列ごとに値を取り出す
            ws.cell(row=data_row_start + row_offset, column=col_idx, value=value)  # セルに値を書き込む


def _write_summary_sheet(ws, summary) -> None:  # サマリ表をExcelに書く
    headers = list(summary.columns)  # サマリの列名を取得する
    for col_idx, name in enumerate(headers, start=1):  # 列名を順番に書き込む
        ws.cell(row=1, column=col_idx, value=name)  # 1行目に列名を配置する

    for row_idx, row in enumerate(summary.itertuples(index=False), start=2):  # 行ごとに値を書き込む
        for col_idx, value in enumerate(row, start=1):  # 列ごとに値を取り出す
            ws.cell(row=row_idx, column=col_idx, value=value)  # セルに値を書き込む


def write_profit_test_excel(path: Path, result: ProfitTestBatchResult) -> Path:  # 収益性検証結果をExcelに書き出す
    """
    Write profit test results to an Excel workbook.
    """
    path.parent.mkdir(parents=True, exist_ok=True)  # 出力先ディレクトリを作成する
    wb = Workbook()  # 新しいExcelブックを作成する
    ws = wb.active  # 既定のシートを取得する
    ws.title = "profit_test"  # シート名を設定する

    first = result.results[0]  # 先頭モデルポイントの結果を代表として使う
    ws["A1"] = "IRR"  # IRRのラベルを設定する
    ws["B1"] = first.irr  # IRRの値を出力する
    ws["A3"] = "New business value"  # NBVのラベルを設定する
    ws["C3"] = first.new_business_value  # NBVの値を出力する

    _write_cashflow_sheet(ws, first)  # キャッシュフロー詳細を出力する

    summary_ws = wb.create_sheet(title="model_point_summary")  # サマリ用シートを作る
    _write_summary_sheet(summary_ws, result.summary)  # サマリ表を出力する

    wb.save(path)  # Excelファイルとして保存する
    return path  # 保存先を返す


def write_profit_test_log(  # 収益性検証のログをテキストで出力する
    path: Path,  # ログ出力先
    config: dict,  # 実行設定
    result: ProfitTestBatchResult,  # 実行結果
) -> Path:  # 保存先を返す
    """
    Write a plain-text log with key assumptions and outputs.
    """
    path.parent.mkdir(parents=True, exist_ok=True)  # 出力先ディレクトリを作成する

    product = config["product"]  # 商品設定を取得する
    pricing = config["pricing"]  # 予定利率や死亡率設定を取得する
    profit_test_cfg = config.get("profit_test", {})  # 収益性検証の設定を取得する
    constraints_cfg = config.get("constraints", {})  # 旧形式の制約設定を取得する
    expense_sufficiency = config.get("expense_sufficiency", {})  # 旧形式の費用充足設定を取得する

    lines = [  # ログの先頭部分を作る
        "profit_test",  # セクション名
        f"term_years(default): {product.get('term_years', 'n/a')}",  # 保険期間のデフォルト
        f"premium_paying_years(default): {product.get('premium_paying_years', 'n/a')}",  # 払込期間のデフォルト
        f"sum_assured(default): {product.get('sum_assured', 'n/a')}",  # 保険金額のデフォルト
        f"pricing_interest_rate: {pricing['interest']['flat_rate']}",  # 予定利率
        f"valuation_interest_rate: {profit_test_cfg.get('valuation_interest_rate', 'default')}",  # 評価利率
        f"lapse_rate: {profit_test_cfg.get('lapse_rate', 'default')}",  # 失効率
        f"irr_min: {constraints_cfg.get('irr_min', 'n/a')}",  # IRR制約（旧設定）
        f"expense_sufficiency: {expense_sufficiency.get('method', 'n/a')}",  # 費用充足の方法（旧設定）
        f"expense_sufficiency_threshold: {expense_sufficiency.get('threshold', 'n/a')}",  # 費用充足の閾値（旧設定）
    ]  # ログのヘッダーここまで

    if result.expense_assumptions is not None:  # 会社費用前提がある場合は詳細を記録する
        lines.extend(  # 会社費用前提の明細を追加する
            [
                "expense_assumptions",  # セクション名
                f"expense_year: {result.expense_assumptions.year}",  # 年度
                f"acq_per_policy: {result.expense_assumptions.acq_per_policy}",  # 獲得費単価
                f"maint_per_policy: {result.expense_assumptions.maint_per_policy}",  # 維持費単価
                f"coll_rate: {result.expense_assumptions.coll_rate}",  # 集金費率
            ]
        )  # 明細の追加

    lines.append("model_point_summary")  # モデルポイントサマリの見出し
    for row in result.summary.itertuples(index=False):  # サマリ行を走査する
        label = row.model_point if hasattr(row, "model_point") else model_point_label(  # ラベル取得の互換処理
            result.results[0].model_point  # model_pointが無い場合は先頭結果を使う
        )  # ラベル決定
        line = (  # 出力行のフォーマットを作る
            f"{label} "  # ラベル
            f"irr={row.irr} "  # IRR
            f"nbv={row.new_business_value} "  # NBV
            f"loading_surplus={row.loading_surplus} "  # 費用充足
            f"premium_to_maturity={row.premium_to_maturity_ratio}"  # PTM比率
        )  # 行の構築
        lines.append(line)  # サマリ行を追加する
        if row.premium_to_maturity_ratio > 1.0:  # 保険料合計が満期保険金を超える場合
            lines.append(f"warning: premium_total_exceeds_maturity {label}")  # 警告を追加する

    path.write_text("\n".join(lines), encoding="utf-8")  # テキストとして保存する
    return path  # 保存先を返す

def write_run_summary_json(
    path: Path,
    config: dict,
    result: ProfitTestBatchResult,
    source: str = "run",
    execution_context: dict[str, object] | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    summary = build_run_summary(
        config,
        result,
        source=source,
        execution_context=execution_context,
    )
    path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")
    return path

def write_optimize_log(  # 最適化結果をテキストで出力する
    path: Path,  # ログ出力先
    config: dict,  # 実行設定
    result: OptimizationResult,  # 最適化結果
) -> Path:  # 保存先を返す
    """
    Write optimization results to a plain-text log.
    """
    path.parent.mkdir(parents=True, exist_ok=True)  # 出力先ディレクトリを作成する

    settings = load_optimization_settings(config)  # 最適化設定を読み込む

    lines = [  # ログのヘッダーを作る
        "optimize",  # セクション名
        f"irr_hard: {settings.irr_hard}",  # IRRハード下限
        f"irr_target: {settings.irr_target}",  # IRRターゲット
        f"loading_surplus_hard: {settings.loading_surplus_hard}",  # 費用充足下限（額）
        f"loading_surplus_hard_ratio: {settings.loading_surplus_hard_ratio}",  # 費用充足下限（比率）
        f"premium_to_maturity_hard_max: {settings.premium_to_maturity_hard_max}",  # PTM上限
        f"premium_to_maturity_target: {settings.premium_to_maturity_target}",  # PTMターゲット
        f"premium_to_maturity_soft_min: {settings.premium_to_maturity_soft_min}",  # PTM soft最小
        f"nbv_hard: {settings.nbv_hard}",  # NBVハード下限
        f"l2_lambda: {settings.l2_lambda}",  # L2重み
        f"objective_mode: {settings.objective_mode}",  # 目的関数モード
        f"min_irr: {result.min_irr}",  # 最小IRR
        f"min_irr_model_point: {result.min_irr_model_point}",  # 最小IRRを与えるモデルポイント
        f"success: {result.success}",  # 成功/失敗
        f"iterations: {result.iterations}",  # 評価回数
        "loading_parameters",  # 係数セクション
        f"a0: {result.params.a0}",  # alpha基礎
        f"a_age: {result.params.a_age}",  # alpha年齢
        f"a_term: {result.params.a_term}",  # alpha期間
        f"a_sex: {result.params.a_sex}",  # alpha性別
        f"b0: {result.params.b0}",  # beta基礎
        f"b_age: {result.params.b_age}",  # beta年齢
        f"b_term: {result.params.b_term}",  # beta期間
        f"b_sex: {result.params.b_sex}",  # beta性別
        f"g0: {result.params.g0}",  # gamma基礎
        f"g_term: {result.params.g_term}",  # gamma期間
    ]  # ログヘッダーここまで

    if result.proposal:  # 条件付き成功の提案がある場合
        changes = result.proposal.get("changes", [])
        change_labels = []
        for change in changes:
            path = change.get("path", "unknown")
            value = change.get("value", "n/a")
            if path == "profit_test.surrender_charge_term":
                change_labels.append(f"Extend surrender_charge_term to {value} years")
            elif path == "optimization.irr_target":
                change_labels.append(f"Lower irr_target to {value}")
            else:
                change_labels.append(f"{path}={value}")
        change_desc = ", ".join(change_labels) if change_labels else "See proposal details"
        lines.append(
            f"WARNING: Default constraints unsatisfied. Proposed Hack: {change_desc}."
        )
        lines.append(f"proposal_plan: {result.proposal.get('plan', 'n/a')}")
        lines.append(
            f"proposal_justification: {result.proposal.get('justification', 'n/a')}"
        )

    if result.watch_model_points is not None:  # 監視対象の指定がある場合
        watch_ids = result.watch_model_points  # 監視対象の一覧
        lines.append(f"watch_list: {', '.join(watch_ids) if watch_ids else 'none'}")  # 監視対象を出力

    if result.exempt_model_points is not None:  # 免除対象の指定がある場合
        exempt_ids = result.exempt_model_points  # 免除対象の一覧
        lines.append(f"exempt_list: {', '.join(exempt_ids) if exempt_ids else 'none'}")  # 免除対象を出力
        if result.exemption_settings is not None:  # 免除設定がある場合は詳細を出す
            sweep = result.exemption_settings.sweep  # sweep設定を取得する
            for model_id in exempt_ids:  # 免除対象ごとに詳細を出力する
                lines.append(  # 詳細行を追加する
                    "exempt_detail "  # 行の接頭辞
                    f"id={model_id} "  # モデルポイントID
                    f"start={sweep.start} "  # 開始値
                    f"end={sweep.end} "  # 終了値
                    f"step={sweep.step} "  # 刻み
                    f"irr_threshold={sweep.irr_threshold}"  # IRR閾値
                )  # 詳細行の追加

    lines.append("model_point_summary")  # モデルポイントサマリの見出し
    for row in result.batch_result.summary.itertuples(index=False):  # サマリ行を走査する
        label = row.model_point  # モデルポイントラベルを取得する
        if label in result.exempt_model_points:  # 免除対象ならステータスだけ出す
            lines.append(  # 免除ステータスの行を追加する
                f"{label} status=exempt"
            )  # 免除ステータスを出力
            continue  # 次の行へ進む
        if label in result.watch_model_points:  # 監視対象の場合は監視として出力する
            threshold = loading_surplus_threshold(settings, int(row.sum_assured))  # 閾値を計算する
            loading_ratio = row.loading_surplus / float(row.sum_assured)  # 比率を計算する
            lines.append(  # 監視の結果行を追加する
                f"{label} irr={row.irr} "  # IRR
                f"nbv={row.new_business_value} "  # NBV
                f"loading_surplus={row.loading_surplus} "  # 充足額
                f"premium_to_maturity={row.premium_to_maturity_ratio} "  # PTM比率
                f"loading_surplus_threshold={threshold} "  # 閾値
                f"loading_surplus_ratio={loading_ratio} "  # 比率
                f"status=watch"  # ステータス
            )  # 行の追加
            if row.premium_to_maturity_ratio > 1.0:  # 保険料合計が満期保険金を超える場合
                lines.append(f"warning: premium_total_exceeds_maturity {label}")  # 警告を追加する
            continue  # 次の行へ進む
        threshold = loading_surplus_threshold(settings, int(row.sum_assured))  # 閾値を計算する
        loading_ratio = row.loading_surplus / float(row.sum_assured)  # 比率を計算する
        irr_shortfall = max(settings.irr_hard - row.irr, 0.0)  # IRRの不足分を計算する
        loading_shortfall = max(threshold - row.loading_surplus, 0.0)  # 充足額不足を計算する
        premium_excess = max(  # PTM上限超過分を計算する
            row.premium_to_maturity_ratio - settings.premium_to_maturity_hard_max, 0.0
        )  # 超過分の計算
        nbv_shortfall = max(settings.nbv_hard - row.new_business_value, 0.0)  # NBV不足を計算する
        status = (  # 4つの制約がすべて満たされるかでステータス判定
            "pass"
            if irr_shortfall <= 0.0
            and loading_shortfall <= 0.0
            and premium_excess <= 0.0
            and nbv_shortfall <= 0.0
            else "fail"
        )  # ステータスの決定
        lines.append(  # 通常の結果行を追加する
            f"{label} irr={row.irr} "  # IRR
            f"nbv={row.new_business_value} "  # NBV
            f"loading_surplus={row.loading_surplus} "  # 充足額
            f"premium_to_maturity={row.premium_to_maturity_ratio} "  # PTM比率
            f"loading_surplus_threshold={threshold} "  # 閾値
            f"loading_surplus_ratio={loading_ratio} "  # 比率
            f"status={status}"  # ステータス
        )  # 行の追加
        if status == "fail":  # 失敗の場合は不足情報を出す
            if irr_shortfall > 0.0:  # IRRが不足している場合
                lines.append(f"shortfall: irr_hard {label} {irr_shortfall:.6f}")  # 不足分を出力
            if loading_shortfall > 0.0:  # 充足額が不足している場合
                lines.append(  # 不足分を出力
                    f"shortfall: loading_surplus_hard {label} {loading_shortfall:.2f}"
                )  # 不足分の出力
            if premium_excess > 0.0:  # PTM上限超過の場合
                lines.append(  # 超過分を出力
                    f"shortfall: premium_to_maturity_hard {label} {premium_excess:.6f}"
                )  # 超過分の出力
            if nbv_shortfall > 0.0:  # NBV不足の場合
                lines.append(f"shortfall: nbv_hard {label} {nbv_shortfall:.2f}")  # 不足分を出力
        if row.premium_to_maturity_ratio > 1.0:  # 保険料合計が満期保険金を超える場合
            lines.append(f"warning: premium_total_exceeds_maturity {label}")  # 警告を追加する

    if result.failure_details:  # 最適化内部で収集した失敗詳細があれば出力する
        lines.append("constraint_failures")  # セクション見出しを追加する
        lines.extend(result.failure_details)  # 失敗詳細を追加する

    path.write_text("\n".join(lines), encoding="utf-8")  # テキストとして保存する
    return path  # 保存先を返す
