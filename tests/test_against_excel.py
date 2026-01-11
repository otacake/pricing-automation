from __future__ import annotations  # 型注釈の前方参照を許可し、循環参照を避けるため

import math  # 近似比較に使うため
import subprocess  # 外部スクリプトの実行確認に使うため
import sys  # 実行パスと標準エラー出力に使うため
from dataclasses import dataclass  # テーブル構造の定義に使うため
from pathlib import Path  # パス操作をOS非依存で行うため

REPO_ROOT = Path(__file__).resolve().parents[1]  # リポジトリルートを取得する
SRC_ROOT = REPO_ROOT / "src"  # src配下をモジュール探索対象にする
if str(SRC_ROOT) not in sys.path:  # まだ追加されていない場合
    sys.path.insert(0, str(SRC_ROOT))  # 先頭に追加して優先度を上げる

import openpyxl  # Excelファイル読み込みに使うため
import pytest  # テスト実行とskipに使うため
import yaml  # YAML設定を読み込むため

from pricing.endowment import calc_endowment_premiums  # Excel比較の対象関数
from pricing.profit_test import run_profit_test  # 収益性検証の検証に使うため

EXCEL_PATH = REPO_ROOT / "data" / "golden" / "養老保険_収益性_RORC.xlsx"  # 参照するExcelファイル


def test_bootstrap_missing_excel() -> None:  # Excelが無い時のエラー動作を検証する
    script = REPO_ROOT / "scripts" / "bootstrap_from_excel.py"  # 対象スクリプトのパス
    missing = REPO_ROOT / "data" / "golden" / "__missing__.xlsx"  # 存在しないパス
    result = subprocess.run(  # スクリプトを実行する
        [sys.executable, str(script), "--xlsx", str(missing)],  # コマンドライン引数
        capture_output=True,  # 標準出力/エラーを取得する
        text=True,  # テキストで受け取る
        check=False,  # 終了コード非ゼロを許容する
    )  # 実行結果
    assert result.returncode == 2  # 終了コードが2であることを確認する
    assert f"File not found: {missing}" in result.stderr  # エラーメッセージを確認する


@dataclass(frozen=True)  # 死亡率テーブルの列位置を保持するため
class _MortalityGroup:  # Excelシート内の死亡率テーブルの位置を表す
    header_row: int  # ヘッダー行
    age_col: int  # 年齢列
    male_col: int  # 男性列
    female_col: int | None  # 女性列


def _coerce_float(value: object) -> float | None:  # Excelセルの値をfloatに変換する
    if value is None or isinstance(value, bool):  # Noneやboolは無効値
        return None  # 無効値としてNoneを返す
    if isinstance(value, (int, float)):  # 数値ならそのまま
        return float(value)  # floatに揃えて返す
    if isinstance(value, str):  # 文字列なら数値化を試す
        text = value.strip()  # 空白を除去する
        if not text:  # 空文字は無効
            return None  # 無効値としてNoneを返す
        try:  # 文字列の数値変換を試す
            return float(text)  # 変換成功値を返す
        except ValueError:  # 変換失敗時
            return None  # 無効値としてNoneを返す
    return None  # その他の型は対象外


def _coerce_int(value: object) -> int | None:  # Excelセルの値をintに変換する
    num = _coerce_float(value)  # まずfloat化する
    if num is None:  # 無効ならNone
        return None  # Noneを返す
    return int(round(num))  # 四捨五入してintにする


def _is_label(value: object, keyword: str) -> bool:  # ラベル文字列を判定する
    if not isinstance(value, str):  # 文字列でないなら対象外
        return False  # Falseを返す
    return keyword in value.replace(" ", "").replace("　", "")  # 空白除去後にキーワードを含むか判定する


def _find_mortality_groups(ws) -> list[_MortalityGroup]:  # 死亡率テーブルの位置を探索する
    groups: list[_MortalityGroup] = []  # 検出結果を初期化する
    max_rows = min(ws.max_row, 50)  # 探索範囲を制限する
    max_cols = min(ws.max_column, 30)  # 探索範囲を制限する
    for row in range(1, max_rows + 1):  # 行方向に探索する
        for col in range(1, max_cols + 1):  # 列方向に探索する
            if not _is_label(ws.cell(row, col).value, "年齢"):  # 年齢ラベルが無ければスキップ
                continue  # 次のセルへ
            if not _is_label(ws.cell(row, col + 1).value, "男性"):  # 男性ラベルが無ければスキップ
                continue  # 次のセルへ
            female_col = None  # 女性列の初期値
            if _is_label(ws.cell(row, col + 2).value, "女性"):  # 女性ラベルがあれば列を設定する
                female_col = col + 2  # 女性列の位置
            groups.append(_MortalityGroup(row, col, col + 1, female_col))  # グループを追加する
    return sorted(groups, key=lambda g: g.age_col)  # 年齢列の位置で並べ替える


def _extract_mortality_rows(ws, group: _MortalityGroup) -> list[dict[str, float | int | None]]:  # 死亡率行を抽出する
    rows: list[dict[str, float | int | None]] = []  # 結果行を初期化する
    start_row = group.header_row + 1  # ヘッダーの次行から開始する
    started = False  # データ開始のフラグ
    for row in range(start_row, ws.max_row + 1):  # 行を走査する
        male = _coerce_float(ws.cell(row, group.male_col).value)  # 男性死亡率を取得する
        female = None  # 女性死亡率の初期値
        if group.female_col is not None:  # 女性列があれば取得する
            female = _coerce_float(ws.cell(row, group.female_col).value)  # 女性死亡率を取得する
        if male is None and female is None:  # 両方欠損なら終端判定
            if started:  # 既に開始していれば終了
                break  # ループを抜ける
            continue  # まだ開始前ならスキップする
        started = True  # データ開始済みを記録する
        age = _coerce_int(ws.cell(row, group.age_col).value)  # 年齢を取得する
        if age is None:  # 年齢が欠損なら行番号から推定する
            age = row - start_row  # 連番として扱う
        rows.append({"age": age, "q_male": male, "q_female": female})  # 行を追加する
    return rows  # 抽出結果を返す


def _require_int(value: object, label: str) -> int:  # 必須整数値を取り出す
    parsed = _coerce_int(value)  # 数値変換を試す
    if parsed is None:  # 変換できない場合
        raise AssertionError(f"Missing {label}")  # テスト失敗として扱う
    return parsed  # 変換値を返す


def _require_float(value: object, label: str) -> float:  # 必須浮動小数値を取り出す
    parsed = _coerce_float(value)  # 数値変換を試す
    if parsed is None:  # 変換できない場合
        raise AssertionError(f"Missing {label}")  # テスト失敗として扱う
    return parsed  # 変換値を返す


def _load_workbook_or_skip():  # Excelが無ければテストをスキップする
    if not EXCEL_PATH.is_file():  # Excelが存在しない場合
        pytest.skip(f"Excel not found: {EXCEL_PATH}")  # スキップする
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)  # Excelを読み込んで返す


def _get_sheet(wb, title: str):  # シートを名前で取得する
    for name in wb.sheetnames:  # シート名を走査する
        if name == title:  # 目的の名前があれば
            return wb[name]  # シートを返す
    raise KeyError(f"Worksheet {title} not found.")  # 見つからなければエラー


def _sex_from_master(value: object) -> str:  # Excelの性別表現をmale/femaleに変換する
    if isinstance(value, str):  # 文字列の場合
        text = value.strip().lower()  # 小文字化して正規化する
        if text in {"male", "m", "男"}:  # 男性表現
            return "male"  # maleを返す
        if text in {"female", "f", "女"}:  # 女性表現
            return "female"  # femaleを返す
    if isinstance(value, (int, float)):  # 数値の場合
        return "male" if int(round(value)) == 1 else "female"  # 1をmale、それ以外をfemaleとする
    return "male"  # 不明な場合はmaleに寄せる


def test_endowment_against_excel() -> None:  # 保険料計算がExcelと一致するか検証する
    wb = _load_workbook_or_skip()  # Excelを読み込む
    try:  # 読み込み後の処理を実行する
        ws_master = _get_sheet(wb, "マスタ")  # マスタシート
        ws_mortality = _get_sheet(wb, "死亡率")  # 死亡率シート

        expected_A = _require_float(ws_master["F2"].value, "A")  # Aの期待値
        expected_a = _require_float(ws_master["F3"].value, "a")  # aの期待値
        expected_net = _require_int(ws_master["G3"].value, "net annual premium")  # 純保険料の期待値
        expected_gross = _require_int(ws_master["F6"].value, "gross annual premium")  # 総保険料の期待値
        expected_monthly = _require_int(ws_master["F7"].value, "monthly premium")  # 月払の期待値

        issue_age = _require_int(ws_master["C2"].value, "issue age")  # 年齢
        sex = _sex_from_master(ws_master["C3"].value)  # 性別
        term_years = _require_int(ws_master["C4"].value, "term years")  # 期間
        premium_paying_years = _require_int(ws_master["C5"].value, "premium paying years")  # 払込期間
        sum_assured = _require_int(ws_master["C6"].value, "sum assured")  # 保険金額
        interest_rate = _require_float(ws_master["C8"].value, "interest rate")  # 利率
        alpha = _require_float(ws_master["C14"].value, "alpha")  # alpha
        beta = _require_float(ws_master["C15"].value, "beta")  # beta
        gamma = _require_float(ws_master["C16"].value, "gamma")  # gamma

        groups = _find_mortality_groups(ws_mortality)  # 死亡率テーブルの位置を探す
        assert groups, "Mortality headers not found."  # 見つからない場合は失敗
        pricing_group = groups[0]  # 予定死亡率テーブルを選ぶ
        mortality_rows = _extract_mortality_rows(ws_mortality, pricing_group)  # 予定死亡率を抽出する

        premiums = calc_endowment_premiums(  # 保険料を計算する
            mortality_rows=mortality_rows,  # 予定死亡率
            sex=sex,  # 性別
            issue_age=issue_age,  # 年齢
            term_years=term_years,  # 期間
            premium_paying_years=premium_paying_years,  # 払込期間
            interest_rate=interest_rate,  # 利率
            sum_assured=sum_assured,  # 保険金額
            alpha=alpha,  # alpha
            beta=beta,  # beta
            gamma=gamma,  # gamma
        )  # 計算結果

        assert math.isclose(premiums.A, expected_A, abs_tol=1e-6)  # Aが一致することを確認する
        assert math.isclose(premiums.a, expected_a, abs_tol=1e-6)  # aが一致することを確認する
        assert premiums.net_annual_premium == expected_net  # 純保険料が一致することを確認する
        assert premiums.gross_annual_premium == expected_gross  # 総保険料が一致することを確認する
        assert premiums.monthly_premium == expected_monthly  # 月払が一致することを確認する
    finally:  # 後処理としてワークブックを閉じる
        wb.close()  # ワークブックを閉じる


def test_profit_test_against_excel() -> None:  # 収益性検証がExcelと一致するか検証する
    wb = _load_workbook_or_skip()  # Excelを読み込む
    try:  # 読み込み後の処理を実行する
        ws_profit = _get_sheet(wb, "収益性検証")  # 収益性検証シート
        ws_master = _get_sheet(wb, "マスタ")  # マスタシート
        expected_irr = _require_float(ws_profit["B1"].value, "IRR")  # IRRの期待値
        expected_nbv = _require_float(ws_profit["C3"].value, "new business value")  # NBVの期待値
        issue_age = _require_int(ws_master["C2"].value, "issue age")  # 年齢
        sex = _sex_from_master(ws_master["C3"].value)  # 性別
        term_years = _require_int(ws_master["C4"].value, "term years")  # 期間
        premium_paying_years = _require_int(ws_master["C5"].value, "premium paying years")  # 払込期間
        sum_assured = _require_int(ws_master["C6"].value, "sum assured")  # 保険金額
    finally:  # 後処理としてワークブックを閉じる
        wb.close()  # ワークブックを閉じる

    config_path = REPO_ROOT / "configs" / "trial-001.yaml"  # 設定ファイルのパス
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))  # 設定を読み込む
    config.pop("model_points", None)  # 複数定義を削除する
    config["model_point"] = {  # 単独モデルポイントとして設定する
        "issue_age": issue_age,  # 年齢
        "sex": sex,  # 性別
        "term_years": term_years,  # 期間
        "premium_paying_years": premium_paying_years,  # 払込期間
        "sum_assured": sum_assured,  # 保険金額
    }  # モデルポイント設定
    profit_test_cfg = dict(config.get("profit_test", {}))  # 収益性検証設定をコピーする
    profit_test_cfg["expense_model"] = {"mode": "loading"}  # Excel比較のため費用モデルをloadingに切り替える
    config["profit_test"] = profit_test_cfg  # 設定を更新する
    result = run_profit_test(config, base_dir=REPO_ROOT)  # 収益性検証を実行する
    single = result.results[0]  # 単一モデルポイントの結果を取得する

    assert math.isclose(single.irr, expected_irr, abs_tol=1e-9)  # IRRが一致することを確認する
    assert math.isclose(single.new_business_value, expected_nbv, abs_tol=1e-6)  # NBVが一致することを確認する
