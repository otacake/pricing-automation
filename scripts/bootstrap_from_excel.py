from __future__ import annotations  # 型注釈の前方参照を許可し、循環参照を避けるため

import argparse  # CLI引数を扱うため
import csv  # CSV書き込みに使うため
import math  # NaN判定や丸めに使うため
import sys  # 標準エラー出力や終了コードに使うため
from dataclasses import dataclass  # 構造化された入力を定義するため
from pathlib import Path  # パス操作をOS非依存で行うため

import openpyxl  # Excelファイル読み込みに使うため

REPO_ROOT = Path(__file__).resolve().parents[1]  # リポジトリルートを基準パスとして使うため
DEFAULT_XLSX = REPO_ROOT / "data" / "golden" / "養老保険_収益性_RORC.xlsx"  # 既定のExcelパス
OUTPUT_DIR = REPO_ROOT / "data"  # 出力先ディレクトリ


@dataclass(frozen=True)  # 死亡率の表構造を不変で保持するため
class MortalityGroup:  # 死亡率テーブルの列位置をまとめる
    header_row: int  # ヘッダー行番号
    age_col: int  # 年齢列
    male_col: int  # 男性死亡率列
    female_col: int | None  # 女性死亡率列（無い場合あり）


def parse_args() -> argparse.Namespace:  # 引数解析を関数化して読みやすくする
    parser = argparse.ArgumentParser(  # 引数パーサを作成する
        description="Extract CSV inputs from the golden Excel file."
    )  # 説明文を設定する
    parser.add_argument(  # Excelパス引数を追加する
        "--xlsx",  # 引数名
        default=str(DEFAULT_XLSX),  # 既定のパス
        help="Path to the Excel file (default: data/golden/養老保険_収益性_RORC.xlsx).",  # 説明文
    )  # 引数定義
    return parser.parse_args()  # 解析結果を返す


def coerce_number(value: object) -> float | None:  # Excelセルの値を数値に変換する
    if value is None or isinstance(value, bool):  # Noneやboolは無効扱い
        return None  # 変換できないのでNoneを返す
    if isinstance(value, (int, float)):  # 既に数値ならそのまま
        if isinstance(value, float) and math.isnan(value):  # NaNは無効
            return None  # NaNは除外する
        return float(value)  # floatに揃えて返す
    if isinstance(value, str):  # 文字列は数値化を試みる
        text = value.strip()  # 前後の空白を除去する
        if not text or text.startswith("="):  # 空文字や数式は無視する
            return None  # 無効値としてNoneを返す
        try:  # 文字列の数値変換を試す
            return float(text.replace(",", ""))  # カンマ除去してfloat化する
        except ValueError:  # 変換失敗時
            return None  # 無効値としてNoneを返す
    return None  # その他の型は対象外とする


def format_int(value: int | float | None) -> str:  # 整数としてCSVに書き出すための整形
    if value is None or isinstance(value, bool):  # 無効値は空文字にする
        return ""  # 空文字を返す
    try:  # 数値変換を試みる
        return str(int(round(float(value))))  # 四捨五入して整数文字列にする
    except (TypeError, ValueError):  # 変換できない場合
        return ""  # 空文字を返す


def format_float(value: int | float | None) -> str:  # 浮動小数としてCSVに書き出すための整形
    if value is None or isinstance(value, bool):  # 無効値は空文字にする
        return ""  # 空文字を返す
    try:  # 数値変換を試みる
        number = float(value)  # floatに変換する
    except (TypeError, ValueError):  # 変換できない場合
        return ""  # 空文字を返す
    if math.isnan(number):  # NaNは無効
        return ""  # 空文字を返す
    text = f"{number:.15f}".rstrip("0").rstrip(".")  # 末尾の不要な0を削除する
    return text if text else "0"  # 空なら0を返す


def is_label(value: object, keyword: str) -> bool:  # ラベル文字列を検出する
    if not isinstance(value, str):  # 文字列でなければ対象外
        return False  # ラベルではない
    return keyword in value.replace(" ", "")  # 空白除去後にキーワードを含むか判定する


def find_mortality_groups(ws) -> list[MortalityGroup]:  # 死亡率テーブルの位置を探索する
    groups: list[MortalityGroup] = []  # 検出結果を初期化する
    max_rows = min(50, ws.max_row)  # 探索範囲を制限する
    max_cols = min(50, ws.max_column)  # 探索範囲を制限する
    for row in range(1, max_rows + 1):  # 行方向に探索する
        for col in range(1, max_cols + 1):  # 列方向に探索する
            if not is_label(ws.cell(row, col).value, "年齢"):  # 年齢ラベルが無ければスキップ
                continue  # 次のセルへ
            if not is_label(ws.cell(row, col + 1).value, "男性"):  # 男性ラベルが無ければスキップ
                continue  # 次のセルへ
            female_col = None  # 女性列は初期値Noneとする
            if is_label(ws.cell(row, col + 2).value, "女性"):  # 女性ラベルがあれば列を設定する
                female_col = col + 2  # 女性列の位置を記録する
            groups.append(  # グループを追加する
                MortalityGroup(  # グループ定義を構築する
                    header_row=row,  # ヘッダー行
                    age_col=col,  # 年齢列
                    male_col=col + 1,  # 男性列
                    female_col=female_col,  # 女性列
                )  # グループの構築
            )  # リストに追加
    groups.sort(key=lambda g: g.age_col)  # 左から順に並べて価格/実績を区別しやすくする
    return groups  # グループ一覧を返す


def extract_mortality_rows(ws, group: MortalityGroup) -> list[tuple[int, float | None, float | None]]:  # 死亡率行を抽出する
    start_row = group.header_row + 1  # ヘッダーの次行から開始する
    rows: list[tuple[int, float | None, float | None]] = []  # 結果行を初期化する
    started = False  # データ開始のフラグ
    for row in range(start_row, ws.max_row + 1):  # 行を走査する
        male_value = coerce_number(ws.cell(row, group.male_col).value)  # 男性死亡率を取得する
        female_value = None  # 女性死亡率の初期値
        if group.female_col is not None:  # 女性列が存在する場合
            female_value = coerce_number(ws.cell(row, group.female_col).value)  # 女性死亡率を取得する
        if male_value is None and female_value is None:  # 両方欠損なら終端判定
            if started:  # 既に開始していれば終了する
                break  # ループを抜ける
            continue  # まだ開始前ならスキップする
        started = True  # データ開始済みを記録する
        age_value = coerce_number(ws.cell(row, group.age_col).value)  # 年齢を取得する
        if age_value is None:  # 年齢が欠損なら行番号から推定する
            age_value = row - start_row  # 連番として扱う
        rows.append((int(round(age_value)), male_value, female_value))  # 行を追加する
    return rows  # 抽出結果を返す


def extract_spot_curve(ws, spot_col: int = 18) -> list[tuple[int, float]]:  # スポットカーブを抽出する
    rows: list[tuple[int, float]] = []  # 結果行を初期化する
    started = False  # データ開始のフラグ
    t = 1  # 期間カウンタの初期値
    for row in range(1, ws.max_row + 1):  # 行を走査する
        value = coerce_number(ws.cell(row, spot_col).value)  # スポット値を取得する
        if value is None:  # 欠損の場合
            if started:  # 既に開始していれば終了
                break  # ループを抜ける
            continue  # 開始前はスキップする
        started = True  # データ開始済みを記録する
        rows.append((t, value / 100.0))  # %表示を小数に変換して追加する
        t += 1  # 期間を増やす
    return rows  # 抽出結果を返す


def write_csv(path: Path, header: list[str], rows: list[list[str]]) -> int:  # CSVを書き出す
    path.parent.mkdir(parents=True, exist_ok=True)  # 出力先ディレクトリを作る
    with path.open("w", encoding="utf-8", newline="") as handle:  # UTF-8でCSVを開く
        writer = csv.writer(handle)  # CSVライターを作る
        writer.writerow(header)  # ヘッダーを書き込む
        writer.writerows(rows)  # データ行を書き込む
    return len(rows) + 1  # 書き込んだ行数（ヘッダー含む）を返す


def print_preview(path: Path, row_count: int, max_lines: int = 5) -> None:  # 書き出し結果のプレビューを表示する
    display_path = path  # 表示用パスを初期化する
    try:  # リポジトリ相対に変換できるか試す
        display_path = path.relative_to(REPO_ROOT)  # 相対パスに変換する
    except ValueError:  # 相対化できない場合はそのまま使う
        pass  # そのまま表示する
    print(str(display_path))  # パスを表示する
    print(f"rows: {row_count}")  # 行数を表示する
    with path.open("r", encoding="utf-8") as handle:  # CSVを開いて先頭を読む
        for _ in range(max_lines):  # 指定行数だけ表示する
            line = handle.readline()  # 1行読む
            if not line:  # EOFなら終了
                break  # ループを抜ける
            print(line.rstrip("\r\n"))  # 改行を除去して表示する
    print("")  # 空行で区切る


def main() -> int:  # スクリプトのメイン処理をまとめる
    args = parse_args()  # 引数を解析する
    xlsx_path = Path(args.xlsx)  # パスをPathに変換する
    if not xlsx_path.is_file():  # ファイルが存在しない場合
        print(f"File not found: {xlsx_path}", file=sys.stderr)  # 標準エラーに出力する
        return 2  # エラー終了コードを返す

    try:  # Excelの読み込みとシート取得を試す
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)  # 数式結果で読み込む
        mortality_ws = workbook["死亡率"]  # 死亡率シート
        spot_ws = workbook["収益性検証_基礎率"]  # スポットカーブシート
    except Exception as exc:  # 読み込み失敗時
        print(exc, file=sys.stderr)  # エラー内容を表示する
        return 3  # エラー終了コードを返す

    try:  # 抽出処理を実行する
        groups = find_mortality_groups(mortality_ws)  # 死亡率テーブル位置を探す
        if not groups:  # 見つからない場合
            raise ValueError("Mortality headers not found.")  # エラーを発生させる
        pricing_group = groups[0]  # 予定死亡率グループを選ぶ
        actual_group = groups[1] if len(groups) > 1 else None  # 実績死亡率グループを選ぶ
        if actual_group is None:  # 実績が無ければエラー
            raise ValueError("Actual mortality headers not found.")  # エラーを発生させる

        pricing_rows_raw = extract_mortality_rows(mortality_ws, pricing_group)  # 予定死亡率を抽出する
        actual_rows_raw = extract_mortality_rows(mortality_ws, actual_group)  # 実績死亡率を抽出する
        spot_rows_raw = extract_spot_curve(spot_ws)  # スポットカーブを抽出する

        pricing_rows = [  # CSV用に整形する
            [format_int(age), format_float(male), format_float(female)]
            for age, male, female in pricing_rows_raw
        ]
        actual_rows = [  # CSV用に整形する
            [format_int(age), format_float(male), format_float(female)]
            for age, male, female in actual_rows_raw
        ]
        spot_rows = [  # CSV用に整形する
            [format_int(t), format_float(rate)] for t, rate in spot_rows_raw
        ]
    except Exception as exc:  # 抽出中のエラー処理
        print(exc, file=sys.stderr)  # エラー内容を表示する
        return 3  # エラー終了コードを返す
    finally:  # 後処理としてワークブックを閉じる
        try:  # closeが失敗しても落とさないためのガード
            workbook.close()  # ワークブックを閉じる
        except Exception:  # close失敗時は無視する
            pass  # 例外を無視する

    outputs = [  # 出力対象の定義
        (OUTPUT_DIR / "mortality_pricing.csv", ["age", "q_male", "q_female"], pricing_rows),  # 予定死亡率
        (OUTPUT_DIR / "mortality_actual.csv", ["age", "q_male", "q_female"], actual_rows),  # 実績死亡率
        (OUTPUT_DIR / "spot_curve_actual.csv", ["t", "spot_rate"], spot_rows),  # スポットカーブ
    ]  # 出力対象の一覧

    for path, header, rows in outputs:  # 出力対象を順に書き出す
        row_count = write_csv(path, header, rows)  # CSVを書き出す
        print_preview(path, row_count)  # 書き出し内容をプレビューする

    return 0  # 正常終了コードを返す


if __name__ == "__main__":  # 直接実行時のみmainを呼ぶ
    sys.exit(main())  # 終了コードを返す
